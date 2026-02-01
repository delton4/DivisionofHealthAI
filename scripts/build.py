#!/usr/bin/env python3
import json
import os
import re
import shutil
import zipfile
import posixpath
from pathlib import Path, PurePosixPath
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader, select_autoescape

ROOT = Path(__file__).resolve().parents[1]
DATA_XLSX = ROOT / "DataForSite.xlsx"
DATA_DIR = ROOT / "data"
DIST_DIR = ROOT / "dist"
TEMPLATES_DIR = ROOT / "templates"
ASSETS_DIR = ROOT / "assets"

BASE_URL = os.environ.get("BASE_URL", "").strip()
if BASE_URL:
    BASE_URL = "/" + BASE_URL.strip("/") + "/"
else:
    BASE_URL = ""

LIST_FIELDS = {"projectIds", "publicationIds", "researcherIds"}

HEADER_MAPS = {
    "Researchers": {
        "researcherid": "id",
        "researchername": "name",
        "researchertitle": "title",
        "researcherabout": "about",
        "researcherprojectids": "projectIds",
        "researcherpublicationid": "publicationIds",
        "researcherimage": "image",
    },
    "Projects": {
        "projectid": "id",
        "projectname": "name",
        "projecttitle": "title",
        "projectabout": "about",
        "projectresearcherids": "researcherIds",
        "projectpublicationid": "publicationIds",
        "projectimage": "image",
        "projectpillar": "pillar",
    },
    "Publications": {
        "publicatoinid": "id",  # typo in sheet
        "publicationid": "id",
        "publicationname": "name",
        "publicationjournal": "journal",
        "publicationabstract": "abstract",
        "publicationurl": "publicationUrl",
        "publicationresearcherids": "researcherIds",
        "publicationprojectid": "projectIds",
        "publicationimage": "image",
    },
}

REQUIRED_FIELDS = {"id", "name"}


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def cell_to_str(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if text.startswith("#") and text.endswith("!"):
        return ""
    return text

def normalize_image_value(value):
    raw = cell_to_str(value)
    if not raw:
        return ""
    lowered = raw.lower()
    if lowered.startswith("http://") or lowered.startswith("https://") or lowered.startswith("//") or lowered.startswith("data:"):
        return raw
    cleaned = raw.strip()
    while cleaned.startswith("./"):
        cleaned = cleaned[2:]
    cleaned = cleaned.lstrip("/")
    if "/" not in cleaned:
        cleaned = f"assets/images/{cleaned}"
    return cleaned


def parse_id_list(value):
    text = cell_to_str(value)
    if not text:
        return []
    parts = re.split(r"[;,\n]+", text)
    cleaned = [p.strip() for p in parts if p and p.strip()]
    seen = set()
    unique = []
    for item in cleaned:
        if item not in seen:
            seen.add(item)
            unique.append(item)
    return unique


def slugify(text):
    base = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return base or "item"


def load_sheet(workbook, sheet_name, header_map, errors):
    if sheet_name not in workbook.sheetnames:
        errors.append({
            "type": "missing_sheet",
            "sheet": sheet_name,
            "message": f"Sheet '{sheet_name}' not found in DataForSite.xlsx",
        })
        return []

    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[0]
    header_index = {}
    image_col_index = None
    for idx, raw in enumerate(header_row):
        normalized = normalize_header(raw)
        if normalized in header_map:
            field = header_map[normalized]
            header_index[idx] = field
            if field == "image":
                image_col_index = idx

    items = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(cell_to_str(value) for value in row):
            continue
        item = {}
        for idx, value in enumerate(row):
            if idx not in header_index:
                continue
            field = header_index[idx]
            if field in LIST_FIELDS:
                item[field] = parse_id_list(value)
            elif field == "image":
                item[field] = normalize_image_value(value)
            elif field == "pillar":
                # Normalize pillar to uppercase
                pillar_value = cell_to_str(value).strip().upper()
                item[field] = pillar_value
            else:
                item[field] = cell_to_str(value)
        # Ensure missing fields are present
        for field in LIST_FIELDS:
            item.setdefault(field, [])
        for field in ["id", "name", "title", "about", "journal", "abstract", "image", "publicationUrl", "pillar"]:
            item.setdefault(field, "")
        item["_row"] = row_number
        items.append(item)

    return items, {"image_col": image_col_index}


def validate_items(items, label, errors):
    seen = {}
    for item in items:
        item_id = item.get("id", "")
        if not item_id:
            errors.append({
                "type": "missing_id",
                "sheet": label,
                "message": f"Missing ID in {label}",
            })
            continue
        if item_id in seen:
            errors.append({
                "type": "duplicate_id",
                "sheet": label,
                "id": item_id,
                "message": f"Duplicate ID '{item_id}' in {label}",
            })
        else:
            seen[item_id] = item

        missing_required = [field for field in REQUIRED_FIELDS if not item.get(field)]
        if missing_required:
            errors.append({
                "type": "missing_required_fields",
                "sheet": label,
                "id": item_id,
                "fields": missing_required,
                "message": f"Missing required fields {missing_required} in {label} ID {item_id}",
            })

    return seen


def validate_references(items, field, target_map, label, target_label, errors):
    for item in items:
        for ref_id in item.get(field, []):
            if ref_id not in target_map:
                errors.append({
                    "type": "missing_reference",
                    "sheet": label,
                    "id": item.get("id", ""),
                    "field": field,
                    "missingId": ref_id,
                    "message": f"Missing {target_label} ID '{ref_id}' referenced by {label} ID {item.get('id', '')}",
                })

def validate_images(items, label, errors, generated_paths=None):
    generated_paths = generated_paths or set()
    for item in items:
        image = item.get("image", "")
        if not image:
            continue
        if image in generated_paths:
            continue
        lowered = image.lower()
        if lowered.startswith("http://") or lowered.startswith("https://") or lowered.startswith("//") or lowered.startswith("data:"):
            continue
        image_path = ROOT / image
        if not image_path.exists():
            errors.append({
                "type": "missing_image",
                "sheet": label,
                "id": item.get("id", ""),
                "image": image,
                "message": f"Image '{image}' not found on disk for {label} ID {item.get('id','')}",
            })

def validate_pillars(items, label, errors):
    """Validate pillar values for projects"""
    valid_pillars = {"PREDICT", "PREVENT", "PERSONALIZE", ""}
    for item in items:
        pillar = item.get("pillar", "")
        if pillar not in valid_pillars:
            errors.append({
                "type": "invalid_pillar",
                "sheet": label,
                "id": item.get("id", ""),
                "pillar": pillar,
                "message": f"Invalid pillar value '{pillar}' in {label} ID {item.get('id','')}. Must be PREDICT, PREVENT, or PERSONALIZE.",
            })

def strip_internal_fields(items):
    for item in items:
        for key in list(item.keys()):
            if key.startswith("_"):
                item.pop(key, None)

def _read_xml(zf, path):
    try:
        with zf.open(path) as handle:
            return ET.parse(handle).getroot()
    except KeyError:
        return None

def _parse_relationships(zf, rels_path):
    rels = {}
    root = _read_xml(zf, rels_path)
    if root is None:
        return rels
    for rel in root:
        rels[rel.attrib.get("Id")] = rel.attrib.get("Target")
    return rels

def _resolve_target(base_path, target):
    if not target:
        return ""
    base_dir = posixpath.dirname(base_path)
    resolved = posixpath.normpath(posixpath.join(base_dir, target))
    return resolved.lstrip("/")

def _get_sheet_paths(zf):
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    workbook = _read_xml(zf, "xl/workbook.xml")
    if workbook is None:
        return {}
    sheets = []
    for sheet in workbook.findall("a:sheets/a:sheet", ns):
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        sheets.append((name, rel_id))
    rels = _parse_relationships(zf, "xl/_rels/workbook.xml.rels")
    mapping = {}
    for name, rel_id in sheets:
        target = rels.get(rel_id, "")
        if not target:
            continue
        sheet_path = "xl/" + target if not target.startswith("xl/") else target
        rels_path = str(PurePosixPath(sheet_path).parent / "_rels" / (PurePosixPath(sheet_path).name + ".rels"))
        mapping[name] = {"path": sheet_path, "rels": rels_path}
    return mapping

def _column_letter_to_index(letters):
    index = 0
    for char in letters.upper():
        if not char.isalpha():
            continue
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(index - 1, 0)

def _cell_ref_to_row_col(cell_ref):
    if not cell_ref:
        return None, None
    match = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper())
    if not match:
        return None, None
    col_letters, row_str = match.groups()
    return int(row_str), _column_letter_to_index(col_letters)

def _build_richvalue_lookup(zf):
    ns_main = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    meta = _read_xml(zf, "xl/metadata.xml")
    if meta is None:
        return {}
    value_meta = meta.find("a:valueMetadata", ns_main)
    if value_meta is None:
        return {}
    vm_to_rv = []
    for bk in value_meta.findall("a:bk", ns_main):
        rc = bk.find("a:rc", ns_main)
        if rc is None:
            continue
        try:
            vm_to_rv.append(int(rc.attrib.get("v", "0")))
        except ValueError:
            vm_to_rv.append(0)

    rv_data = _read_xml(zf, "xl/richData/rdrichvalue.xml")
    if rv_data is None:
        return {}
    ns_rv = {"rd": "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"}
    rv_to_local = []
    for rv in rv_data.findall("rd:rv", ns_rv):
        values = []
        for value in rv.findall("rd:v", ns_rv):
            try:
                values.append(int(value.text))
            except (TypeError, ValueError):
                values.append(None)
        rv_to_local.append(values[0] if values else None)

    rv_rel = _read_xml(zf, "xl/richData/richValueRel.xml")
    if rv_rel is None:
        return {}
    ns_rel = {
        "rv": "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    rids = []
    for rel in rv_rel.findall("rv:rel", ns_rel):
        rid = rel.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if rid:
            rids.append(rid)

    rels = _parse_relationships(zf, "xl/richData/_rels/richValueRel.xml.rels")
    lookup = {}
    for vm_index, rv_index in enumerate(vm_to_rv):
        if rv_index is None or rv_index >= len(rv_to_local):
            continue
        local_id = rv_to_local[rv_index]
        if local_id is None or local_id >= len(rids):
            continue
        rid = rids[local_id]
        target = rels.get(rid)
        if not target:
            continue
        media_path = _resolve_target("xl/richData/richValueRel.xml", target)
        lookup[vm_index] = media_path
    return lookup

def collect_embedded_images(xlsx_path, sheet_name, folder, row_map, image_col_index, errors):
    outputs = []
    generated = set()
    with zipfile.ZipFile(xlsx_path) as zf:
        sheet_paths = _get_sheet_paths(zf)
        if sheet_name not in sheet_paths:
            errors.append({
                "type": "missing_sheet_for_images",
                "sheet": sheet_name,
                "message": f"Sheet '{sheet_name}' not found when extracting images",
            })
            return outputs, generated

        sheet_path = sheet_paths[sheet_name]["path"]
        sheet_rels_path = sheet_paths[sheet_name]["rels"]
        sheet_xml = _read_xml(zf, sheet_path)
        if sheet_xml is None:
            return outputs, generated

        ns_sheet = {
            "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        drawing_ids = []
        for drawing in sheet_xml.findall("a:drawing", ns_sheet):
            drawing_id = drawing.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if drawing_id:
                drawing_ids.append(drawing_id)

        sheet_rels = _parse_relationships(zf, sheet_rels_path)

        for drawing_id in drawing_ids:
            drawing_target = sheet_rels.get(drawing_id)
            if not drawing_target:
                continue
            drawing_path = _resolve_target(sheet_path, drawing_target)
            drawing_xml = _read_xml(zf, drawing_path)
            if drawing_xml is None:
                continue

            drawing_rels_path = str(PurePosixPath(drawing_path).parent / "_rels" / (PurePosixPath(drawing_path).name + ".rels"))
            drawing_rels = _parse_relationships(zf, drawing_rels_path)

            ns_draw = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            }

            anchors = drawing_xml.findall("xdr:twoCellAnchor", ns_draw) + drawing_xml.findall("xdr:oneCellAnchor", ns_draw)
            for anchor in anchors:
                from_node = anchor.find("xdr:from", ns_draw)
                if from_node is None:
                    continue
                row_node = from_node.find("xdr:row", ns_draw)
                col_node = from_node.find("xdr:col", ns_draw)
                if row_node is None or col_node is None:
                    continue
                try:
                    row_index = int(row_node.text)
                    col_index = int(col_node.text)
                except (TypeError, ValueError):
                    continue

                if image_col_index is not None and col_index != image_col_index:
                    continue

                blip = anchor.find(".//a:blip", ns_draw)
                if blip is None:
                    continue
                embed = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                if not embed:
                    continue
                image_target = drawing_rels.get(embed)
                if not image_target:
                    continue
                image_path = _resolve_target(drawing_path, image_target)
                if not image_path:
                    continue

                row_number = row_index + 1
                item = row_map.get(row_number)
                if not item:
                    errors.append({
                        "type": "image_without_row",
                        "sheet": sheet_name,
                        "row": row_number,
                        "message": f"Image anchored on row {row_number} has no matching data row in {sheet_name}",
                    })
                    continue

                if item.get("image"):
                    continue

                item_id = item.get("id", "")
                safe_id = slugify(str(item_id)) if item_id else f"row-{row_number}"
                ext = PurePosixPath(image_path).suffix.lstrip(".").lower() or "png"
                rel_path = f"assets/images/{folder}/{safe_id}.{ext}"
                if rel_path in generated:
                    continue

                try:
                    with zf.open(image_path) as img_handle:
                        img_bytes = img_handle.read()
                except KeyError:
                    errors.append({
                        "type": "missing_embedded_media",
                        "sheet": sheet_name,
                        "row": row_number,
                        "image": image_path,
                        "message": f"Embedded image file '{image_path}' missing for {sheet_name} row {row_number}",
                    })
                    continue

                item["image"] = rel_path
                outputs.append({"rel_path": rel_path, "bytes": img_bytes})
                generated.add(rel_path)

        rich_lookup = _build_richvalue_lookup(zf)
        if rich_lookup:
            for cell in sheet_xml.findall(".//a:c[@vm]", ns_sheet):
                vm_raw = cell.attrib.get("vm")
                if vm_raw is None:
                    continue
                try:
                    vm_index = int(vm_raw)
                except ValueError:
                    continue

                # Convert from 1-indexed (Excel) to 0-indexed (Python dict)
                lookup_key = vm_index - 1
                if lookup_key < 0 or lookup_key >= len(rich_lookup):
                    errors.append({
                        "type": "richvalue_index_out_of_bounds",
                        "sheet": sheet_name,
                        "vm_index": vm_index,
                        "lookup_key": lookup_key,
                        "message": f"Rich value lookup key {lookup_key} (from vm={vm_index}) out of bounds for {sheet_name}",
                    })
                    continue

                image_path = rich_lookup.get(lookup_key)
                if not image_path:
                    errors.append({
                        "type": "richvalue_lookup_failed",
                        "sheet": sheet_name,
                        "vm_index": vm_index,
                        "message": f"Rich value metadata index {vm_index} not found in lookup table for {sheet_name}",
                    })
                    continue

                cell_ref = cell.attrib.get("r")
                row_number, col_index = _cell_ref_to_row_col(cell_ref)
                if row_number is None or col_index is None:
                    continue
                if image_col_index is not None and col_index != image_col_index:
                    continue

                item = row_map.get(row_number)
                if not item:
                    errors.append({
                        "type": "image_without_row",
                        "sheet": sheet_name,
                        "row": row_number,
                        "message": f"Image metadata on row {row_number} has no matching data row in {sheet_name}",
                    })
                    continue
                if item.get("image"):
                    continue

                item_id = item.get("id", "")
                safe_id = slugify(str(item_id)) if item_id else f"row-{row_number}"
                ext = PurePosixPath(image_path).suffix.lstrip(".").lower() or "png"
                rel_path = f"assets/images/{folder}/{safe_id}.{ext}"
                if rel_path in generated:
                    continue
                try:
                    with zf.open(image_path) as img_handle:
                        img_bytes = img_handle.read()
                except KeyError:
                    errors.append({
                        "type": "missing_embedded_media",
                        "sheet": sheet_name,
                        "row": row_number,
                        "image": image_path,
                        "message": f"Embedded image file '{image_path}' missing for {sheet_name} row {row_number}",
                    })
                    continue

                item["image"] = rel_path
                outputs.append({"rel_path": rel_path, "bytes": img_bytes})
                generated.add(rel_path)

    return outputs, generated

def add_meta(items, kind):
    for item in items:
        display = item.get("name") or item.get("title") or item.get("id") or "item"
        item["slug"] = slugify(display)
        item["path"] = f"{kind}/{item.get('id','')}-{item['slug']}.html"


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def build_site(data):
    env = Environment(
        loader=FileSystemLoader(TEMPLATES_DIR),
        autoescape=select_autoescape(["html"]),
    )

    root_prefix = BASE_URL
    sub_prefix = BASE_URL if BASE_URL else "../"

    DIST_DIR.mkdir(parents=True, exist_ok=True)

    # index
    index_template = env.get_template("index.html")
    sections = [
        {"slug": "researchers", "title": "Researchers", "count": len(data["researchers"])},
        {"slug": "projects", "title": "Projects", "count": len(data["projects"])},
        {"slug": "publications", "title": "Publications", "count": len(data["publications"])},
    ]
    index_html = index_template.render(
        base_url=root_prefix,
        page_title="Division of Health AI",
        sections=sections,
    )
    (DIST_DIR / "index.html").write_text(index_html, encoding="utf-8")

    list_template = env.get_template("list.html")
    detail_template = env.get_template("detail.html")

    # list pages
    for kind, title, items in [
        ("researchers", "Researchers", data["researchers"]),
        ("projects", "Projects", data["projects"]),
        ("publications", "Publications", data["publications"]),
    ]:
        section_dir = DIST_DIR / kind
        section_dir.mkdir(parents=True, exist_ok=True)
        list_html = list_template.render(
            base_url=sub_prefix,
            page_title=title,
            page_description=f"All {title.lower()} in the Division of Health AI database.",
            items=items,
        )
        (section_dir / "index.html").write_text(list_html, encoding="utf-8")

    # detail pages
    researcher_map = {item["id"]: item for item in data["researchers"]}
    project_map = {item["id"]: item for item in data["projects"]}
    publication_map = {item["id"]: item for item in data["publications"]}

    def related(items_map, ids):
        return [items_map[i] for i in ids if i in items_map]

    for researcher in data["researchers"]:
        html = detail_template.render(
            base_url=sub_prefix,
            page_title=researcher.get("name") or researcher.get("id"),
            section="researchers",
            entity=researcher,
            related_projects=related(project_map, researcher.get("projectIds", [])),
            related_publications=related(publication_map, researcher.get("publicationIds", [])),
            related_researchers=[],
        )
        target = DIST_DIR / "researchers" / Path(researcher["path"]).name
        target.write_text(html, encoding="utf-8")

    for project in data["projects"]:
        html = detail_template.render(
            base_url=sub_prefix,
            page_title=project.get("name") or project.get("id"),
            section="projects",
            entity=project,
            related_projects=[],
            related_publications=related(publication_map, project.get("publicationIds", [])),
            related_researchers=related(researcher_map, project.get("researcherIds", [])),
        )
        target = DIST_DIR / "projects" / Path(project["path"]).name
        target.write_text(html, encoding="utf-8")

    for publication in data["publications"]:
        html = detail_template.render(
            base_url=sub_prefix,
            page_title=publication.get("name") or publication.get("id"),
            section="publications",
            entity=publication,
            related_projects=related(project_map, publication.get("projectIds", [])),
            related_publications=[],
            related_researchers=related(researcher_map, publication.get("researcherIds", [])),
        )
        target = DIST_DIR / "publications" / Path(publication["path"]).name
        target.write_text(html, encoding="utf-8")



def main():
    errors = []
    workbook = load_workbook(DATA_XLSX, data_only=True)

    researchers, researcher_meta = load_sheet(workbook, "Researchers", HEADER_MAPS["Researchers"], errors)
    projects, project_meta = load_sheet(workbook, "Projects", HEADER_MAPS["Projects"], errors)
    publications, publication_meta = load_sheet(workbook, "Publications", HEADER_MAPS["Publications"], errors)

    validate_items(researchers, "Researchers", errors)
    validate_items(projects, "Projects", errors)
    validate_items(publications, "Publications", errors)

    researchers = [item for item in researchers if item.get("id")]
    projects = [item for item in projects if item.get("id")]
    publications = [item for item in publications if item.get("id")]

    add_meta(researchers, "researchers")
    add_meta(projects, "projects")
    add_meta(publications, "publications")

    researcher_map = {item["id"]: item for item in researchers}
    project_map = {item["id"]: item for item in projects}
    publication_map = {item["id"]: item for item in publications}

    validate_references(researchers, "projectIds", project_map, "Researchers", "Project", errors)
    validate_references(researchers, "publicationIds", publication_map, "Researchers", "Publication", errors)

    validate_references(projects, "researcherIds", researcher_map, "Projects", "Researcher", errors)
    validate_references(projects, "publicationIds", publication_map, "Projects", "Publication", errors)

    validate_references(publications, "researcherIds", researcher_map, "Publications", "Researcher", errors)
    validate_references(publications, "projectIds", project_map, "Publications", "Project", errors)

    # Validate pillar values for projects
    validate_pillars(projects, "Projects", errors)

    image_outputs = []
    generated_paths = set()

    researcher_row_map = {item.get("_row"): item for item in researchers if item.get("_row")}
    project_row_map = {item.get("_row"): item for item in projects if item.get("_row")}
    publication_row_map = {item.get("_row"): item for item in publications if item.get("_row")}

    outputs, generated = collect_embedded_images(
        DATA_XLSX,
        "Researchers",
        "researchers",
        researcher_row_map,
        researcher_meta.get("image_col"),
        errors,
    )
    image_outputs.extend(outputs)
    generated_paths.update(generated)

    outputs, generated = collect_embedded_images(
        DATA_XLSX,
        "Projects",
        "projects",
        project_row_map,
        project_meta.get("image_col"),
        errors,
    )
    image_outputs.extend(outputs)
    generated_paths.update(generated)

    outputs, generated = collect_embedded_images(
        DATA_XLSX,
        "Publications",
        "publications",
        publication_row_map,
        publication_meta.get("image_col"),
        errors,
    )
    image_outputs.extend(outputs)
    generated_paths.update(generated)

    validate_images(researchers, "Researchers", errors, generated_paths)
    validate_images(projects, "Projects", errors, generated_paths)
    validate_images(publications, "Publications", errors, generated_paths)

    strip_internal_fields(researchers)
    strip_internal_fields(projects)
    strip_internal_fields(publications)

    # Write normalized data
    write_json(DATA_DIR / "researchers.json", researchers)
    write_json(DATA_DIR / "projects.json", projects)
    write_json(DATA_DIR / "publications.json", publications)
    write_json(DATA_DIR / "errors.json", errors)

    # Build output
    if DIST_DIR.exists():
        shutil.rmtree(DIST_DIR)
    DIST_DIR.mkdir(parents=True, exist_ok=True)

    if ASSETS_DIR.exists():
        shutil.copytree(ASSETS_DIR, DIST_DIR / "assets")

    for image in image_outputs:
        target = DIST_DIR / image["rel_path"]
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(image["bytes"])

    data = {
        "researchers": researchers,
        "projects": projects,
        "publications": publications,
        "errors": errors,
    }
    build_site(data)

    # Copy data into dist for optional client access
    dist_data = DIST_DIR / "data"
    dist_data.mkdir(parents=True, exist_ok=True)
    for name in ["researchers.json", "projects.json", "publications.json", "errors.json"]:
        shutil.copy2(DATA_DIR / name, dist_data / name)

    print("Build complete")
    if errors:
        print(f"Validation warnings: {len(errors)} (see data/errors.json)")


if __name__ == "__main__":
    main()
